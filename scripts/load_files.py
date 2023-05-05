#!python
"""
Script to load data about births to a postgres database

NOTE
I've removed incorrect headers on these files for a clean load to a data frame.
Source: https://web.archive.org/web/20181107034101/https://www.phila.gov/health/Commissioner/VitalStatistics.html

Before runnin this script, you need to create the database and schemas.
In the database, don't forget CREATE EXTENSION postgis;
Create sdp.sdp and
       sdp.census
(or change the variables below)
"""

import openpyxl
import os
import pandas
import re
from sqlalchemy import create_engine
import subprocess
from tqdm import tqdm
import urllib.request
import zipfile

census_key = '' # TODO add your key here

def get_csvs(file_path):
    return list(filter(lambda f: re.search(r"\.csv$", f), os.listdir(file_path)))

def clean_col(col: str):
    col = col.lower()
    col = col.replace('+', ' and ')
    col = col.replace('<', '')
    col = col.replace(',', '')
    col = col.replace('.', '')
    col = col.replace('-', '_')
    col = col.replace('/', '_')
    col = col.replace('\n', '_')
    col = col.replace('#', '')
    col = col.replace('%', '')
    col = col.replace('  ', '_')
    col = col.strip()
    col = col.replace(' ', '_')
    return col

def load_births_file(births_file, births_path, engine, schema):
    """Load one births file to database"""
    year = births_file[0:4]
    df = pandas.read_csv(os.path.join(births_path, births_file))
    df.insert(0, 'census_year', year)
    df.rename(columns=clean_col, inplace=True)

    table_name = f'births_{year}'
    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists='replace',
    )

def district_enrollment_csv(download_path, year, engine, schema):
    """
    Load district enrollment in csv format
    Works for school year 2019-2020 and going forward
    """
    url_template = "https://cdn.philasd.org/offices/performance/Open_Data/School_Information/Enrollment_Demographics_School/{year}%20Enrollment%20&%20Demographics.csv"

    url = url_template.format(year=year)
    download_file = f'{year}_Enrollment_Demographics.csv'
    download_file_path = os.path.join(download_path, download_file)
    urllib.request.urlretrieve(url, download_file_path)
    table_name = clean_col(f'schools_demog_{year}')
    df = pandas.read_csv(
        download_file_path,
        sep=',',
        )
    df.rename(columns=clean_col, inplace=True)
    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists='replace',
    )

def district_enrollment_xlsx(download_path, year, engine, schema):
    """
    Load district enrollment in intermediary xlsx format
    Works for school year 2018-2019 only
    """
    url_template = 'https://cdn.philasd.org/offices/performance/Open_Data/School_Information/Enrollment_Demographics_School/{year}%20Enrollment%20&%20Demographics.xlsx'
    
    url = url_template.format(year=year)
    download_file = f'{year}_Enrollment_Demographics.xlsx'
    download_file_path = os.path.join(download_path, download_file)
    urllib.request.urlretrieve(url, download_file_path)

    workbook = openpyxl.load_workbook(download_file_path)
    worksheet = workbook['Sheet1']
    data = worksheet.values
    cols = next(data)[0:]
    data = list(data)
    df = pandas.DataFrame(data, columns=cols)

    table_name = clean_col(f'schools_demog_{year}')
    df.rename(columns=clean_col, inplace=True)
    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists='replace',
    )

def district_enrollment_xlsx_legacy(download_path, year, engine, schema):
    """
    Load district enrollment in legacy xlsx format
    Works for school year 2017-2018 and before
    """
    url_template = 'https://cdn.philasd.org/offices/performance/Open_Data/School_Information/Enrollment_Demographics_School/{year}%20Enrollment%20&%20Demographics.xlsx'

    url = url_template.format(year=year)
    download_file = f'{year}_Enrollment_Demographics.xlsx'
    download_file_path = os.path.join(download_path, download_file)
    urllib.request.urlretrieve(url, download_file_path)

    workbook = openpyxl.load_workbook(download_file_path)

    # process sheet(s) we need

    # Ethnicity
    ethnicity = workbook['Ethnicity'].values
    # skip 4 lines
    for i in range(4):
        next(ethnicity)
    subtabs = next(ethnicity)[1:21]
    ethnicity_cols = list(next(ethnicity)[1:21])
    # match subtabs to columns to build something like "Hispanic_percent"
    for label_idx, label in enumerate(subtabs):
        if label:
            for col_idx in range(len(ethnicity_cols)):
                if (col_idx - 1) // 2 == (label_idx) // 2:
                    ethnicity_cols[col_idx] = label.split('\n')[0] + '_' + ethnicity_cols[col_idx]
    ethnicity = [r[1:21] for r in list(ethnicity)]

    df_ethnicity = pandas.DataFrame(ethnicity, columns=ethnicity_cols)
    df_ethnicity.rename(columns=clean_col, inplace=True)
    # if we merge to another sheet, remove duplicate columns to avoid muddled names
    #df_ethnicity = df_ethnicity.drop(columns=['total_enrolled', 'learning_network', 'school_name'])

    # merge both dataframes together so we only load one table
    #df = pandas.merge(something, df_ethnicity, how='left', on=['school_id', 'grade'])
    df = df_ethnicity
    df .insert(0, 'school_year', year)

    table_name = clean_col(f'schools_demog_{year}')
    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists='replace',
    )

def school_list(list_data_path, engine, mode, schema, year, url, sheet):
    """
    Load school-level data
    :param mode: 1 = replace table; != 1 = append table
                 First file drops and recreates table.
                 Subsequent files append to the new table.
    """
    download_file = os.path.basename(url).replace("%20", "_")
    download_file_path = os.path.join(list_data_path, download_file)
    urllib.request.urlretrieve(url, download_file_path)

    # treat xlsx and csv files differently

    if re.search(r"\.csv$", download_file_path):
        df = pandas.read_csv(download_file_path)
        df.insert(0, "school_year", year)
        df.rename(columns=clean_col, inplace=True)

    elif re.search(r"\.xlsx$", download_file_path):

        workbook = openpyxl.load_workbook(download_file_path)
        active_sheet = workbook[sheet].values

        start_column = 0 # begin, index starts at zero
        n_columns = 57   # continue for n columns

        cols = list(next(active_sheet)[start_column:n_columns])
        data = [r[start_column:n_columns] for r in list(active_sheet)]

        df = pandas.DataFrame(data, columns=cols)

        df.rename(columns=clean_col, inplace=True)
        convert_cols = {
            "grade_span": "current_grade_span_served",
        }
        df.rename(columns=convert_cols, inplace=True)

    # at this point, we have a dataframe we can treat the same way regardless of the file

    keep_cols = [
        "school_year",
        "pa_code",
        "ulcs_code",
        "publication_name",
        "admission_type",
        "current_grade_span_served",
        "school_level",
        "governance",
        "school_region_code",
        "school_region",
        "year_closed",
        "gps_location",
        "school_leader_name",
    ]

    df = df.filter(keep_cols, axis=1)
    df["gps_location"] = df.get("gps_location", None)

    if year != "multi":
        # make the school_year column equal the value we passed in
        df["school_year"] = pandas.Series([year for x in range(len(df.index))])

    if_exists = "append"
    if mode == 0:
        if_exists = "replace"

    table_name = clean_col(f'school')
    df.to_sql(
        name=table_name,
        con=engine,
        schema=schema,
        if_exists=if_exists,
    )

def download_unzip(url: str, download_path) -> str:
    """
    Download and unzip a file
    return the full path to the unzipped contents
    """
    download_file = os.path.basename(url)
    download_file_path = os.path.join(download_path, download_file)
    urllib.request.urlretrieve(url, download_file_path)

    # unzip the file
    unzip_dir = os.path.basename(download_file_path).rstrip('.zip')
    unzip_path = os.path.join(download_path, unzip_dir)
    with zipfile.ZipFile(download_file_path, 'r') as f:
        f.extractall(unzip_path)

    return unzip_path

def unzip_score_file(year, file_paths, destination_dir) -> str:
    """
    Validate the score file exists for the given year.
    If the score file exists, return the path to that file.
    """
    # filter scores to just this year's files
    zip_files = list(filter(lambda p: p != p.rstrip(".zip"), file_paths))
    score_files = list(filter(lambda p: p[0:4] == year, zip_files))

    # confirm a single result
    if len(score_files) == 0:
        print(f"No score file found for year {year}")
    elif len(score_files) > 1:
        print(f"More than one score file found for year {year}")
    
    score_file = score_files[0]

    # unzip the file
    unzip_dir = os.path.basename(score_file).rstrip('.zip')
    unzip_path = os.path.join(destination_dir, unzip_dir)
    score_file_path = os.path.join(destination_dir, score_file)
    with zipfile.ZipFile(score_file_path, 'r') as f:
        f.extractall(unzip_path)

    # build the path to the file we need to process
    scores_year = os.path.basename(score_file)[5:9]
    scores_file = f"{scores_year} PSSA Keystone Actual (School_S).xlsx"
    scores_file_path = os.path.join(unzip_path, scores_file)
    
    return scores_file_path

def scores(file_path, engine, mode, schema, year):
    """
    If the given year's file is available to load, load
    """
    scores_year = os.path.basename(file_path)[0:4]
    workbook = openpyxl.load_workbook(file_path)
    if scores_year <= '2017':
        sheet_name = "All Students"
        sheet = workbook[sheet_name].values
        # L6 has our secondary headers
        for i in range(5):
            next(sheet)
        scores_subtabs = next(sheet)[1:19]
        # L7 has our primary headers
        # col 1 is blank, data ends at S=19
        scores_cols = list(next(sheet)[1:19])
        # match subtabs to columns to build something like "Hispanic_percent"
        for label_idx, label in enumerate(scores_subtabs):
            if label:
                for col_idx in range(len(scores_cols)):
                    if (col_idx) // 2 == (label_idx) // 2:
                        scores_cols[col_idx] = label.split('\n')[0] + '_' + scores_cols[col_idx]
        sheet = [row[1:19] for row in list(sheet)]

        df = pandas.DataFrame(sheet, columns=scores_cols)
        df.rename(columns=clean_col, inplace=True)

    elif scores_year >= '2018':
        sheet_name = "Sheet1"
        sheet = workbook[sheet_name].values
        scores_cols = list(next(sheet)[0:19])
        sheet = [row[0:19] for row in list(sheet)]

        df = pandas.DataFrame(sheet, columns=scores_cols)
        df.rename(columns=clean_col, inplace=True)

        scores_field_map = {
            "src_school_id": "school_code",
            "school_id": "school_code",
            "count_below_basic": "level_1_count",
            "percent_below_basic": "level_1_percent",
            "count_basic": "level_2_count",
            "percent_basic": "level_2_percent",
            "count_proficient": "level_3_count",
            "percent_proficient": "level_3_percent",
            "count_advanced": "level_4_count",
            "percent_advanced": "level_4_percent",
            "count_prof_adv": "levels_3_and_4_count",
            "percent_prof_adv": "levels_3_and_4_percent",
        }

        df.rename(columns=scores_field_map, inplace=True)

    # apply to datafram from either origin
    df.insert(0, 'school_year', scores_year)

    if_exists = 'append'
    if mode == 1:
        if_exists = 'replace'
    df.to_sql(
        name="score",
        con=engine,
        schema=schema,
        if_exists=if_exists,
    )

def census(engine, schema: str, table: str, url: str, fields: list):
    """
    Load table using url, using fields
    """

    df = pandas.read_json(url.format(census_key=census_key))
    df.columns = fields
    df.rename(columns=clean_col, inplace=True)
    df = df.drop(axis='index', index=0)

    df.to_sql(
        name=table,
        con=engine,
        schema=schema,
        if_exists='replace',
    )

    # list names of columns
    #list(df.iloc[0].axes[0])

def unzip_shape_file(year, destination_dir) -> str:
    """
    Unzip shape file contents
    return path to the elementary school shapes directory
    """
    shape_file_zip = f"SDP_Catchment_{year}.zip"
    shape_file_path = os.path.join(destination_dir, shape_file_zip)

    # validate the existence of the shape file
    if not os.path.isfile(shape_file_path):
        print(f"File {shape_file_path} does not exist.")
        return None

    # unzip the file
    unzip_dir = os.path.basename(shape_file_path).rstrip('.zip')
    unzip_path = os.path.join(destination_dir, unzip_dir)
    with zipfile.ZipFile(shape_file_path, 'r') as f:
        f.extractall(unzip_path)

    es_shape_file = os.path.join(
        unzip_path,
        f'Catchment_ES_20{year[0:2]}-{year[2:4]}',
        f'Catchment_ES_20{year[0:2]}.shp')
    
    if not os.path.isfile(es_shape_file):
        print(f"File {es_shape_file} does not exist.")
        return None

    return es_shape_file

def local_shapes(file_path, schema, db, user):
    """
    Load a local shapefile
    """
    subprocess.run(
        [
            'ogr2ogr',
            '-f', 'PostgreSQL',
            f'Pg:dbname={db} host=localhost port=5432 user={user}',
            '-lco', f'SCHEMA={schema}',
            '-lco', 'OVERWRITE=YES',
            '-nlt', 'PROMOTE_TO_MULTI',
            '-t_srs', 'EPSG:2272',
            '-lco', 'precision=NO',
            file_path
        ],
        check=True,
    )

def main():
    user = input('user: ')
    #pw = input('password: ')
    db = 'sdp'
    census_schema = 'census'
    sdp_schema = 'sdp'
    engine = create_engine(f"postgresql+psycopg2://{user}@localhost:5432/{db}")

    print("Loading births files...")
    births_path = os.path.join('..',"data", "births")
    births_files = get_csvs(births_path)

    for births_file in tqdm(births_files):
        load_births_file(births_file, births_path, engine, census_schema)

    print("Loading school district demographics files...")
    # All demog data will live in one directory
    demog_data_path = os.path.join('..', 'data', 'schools', 'demog')

    demog_years = [
        '2019-2020',
        '2018-2019',
        '2017-2018',
        '2016-2017',
    ]

    for year in tqdm(demog_years):
        # TODO use the national education data api to get this data.
        if year >= '2019-2020':
            district_enrollment = district_enrollment_csv
        elif year == '2018-2019':
            district_enrollment = district_enrollment_xlsx
        elif year <= '2017-2018':
            district_enrollment = district_enrollment_xlsx_legacy
        else:
            print(f"Demographics in year {year} is not supported.")

        district_enrollment(demog_data_path, year, engine, sdp_schema)

    print("Processing SDP list of schools...")
    list_data_path = os.path.join('..', 'data', 'schools', 'list')

    school_list_configs = [
        {
            "year": "multi",
            "url": "https://cdn.philasd.org/offices/performance/Open_Data/School_Information/School_List/Longitudinal%20School%20List%20(20171128).xlsx",
            "sheet": "Sheet1",
        },
        {
            "year": "2017-2018",
            "url": "https://cdn.philasd.org/offices/performance/Open_Data/School_Information/School_List/2017-2018%20Master%20School%20List%20(20180611).xlsx",
            "sheet": "Master School List",
        },
        {
            "year": "2018-2019",
            "url": "https://cdn.philasd.org/offices/performance/Open_Data/School_Information/School_List/2018-2019%20Master%20School%20List%20(20190510).csv",
            "sheet": None,
        },
        {
            "year": "2019-2020",
            "url": "https://cdn.philasd.org/offices/performance/Open_Data/School_Information/School_List/2019-2020%20Master%20School%20List%20(20201123).csv",
            "sheet": None,
        },
    ]
    for index, school_list_config in tqdm(enumerate(school_list_configs)):
        school_list(list_data_path, engine, index, sdp_schema, **school_list_config)

    print("Processing sdp scores...")

    # download the zip file of all years
    scores_url = "https://cdn.philasd.org/offices/performance/Open_Data/School_Performance/PSSA_Keystone/PSSA_Keystone_All_Years.zip"
    scores_path = os.path.join('..', 'data', 'schools')

    scores_unzip_path = download_unzip(scores_url, scores_path)

    # build our list of files to process
    scores_all_files = os.listdir(scores_unzip_path)
    scores_zip_files = list(filter(lambda p: p != p.rstrip(".zip"), scores_all_files))

    # the sequence seems to matter here
    score_years = [
        '2017',
        '2018',
        '2016',
        '2015',
    ]

    for index, year in tqdm(enumerate(score_years)):
        # get the path to an unzipped file
        score_file_path = unzip_score_file(year, scores_zip_files, scores_unzip_path)
        # load that unzipped file
        scores(score_file_path, engine, index, sdp_schema, year)

    print("Process block group data...")
    """
    Documentation on census 2010 race columns:
        https://api.census.gov/data/2010/dec/sf1/groups/P3.html
    Documentation on acs 2015 race columns:
        https://www2.census.gov/programs-surveys/acs/summary_file/2015/documentation/user_tools/ACS2015_Table_Shells.xlsx
    """
    census_configs = [
        {
            "table": "census_2010",
            "url": "https://api.census.gov/data/2010/dec/sf1?get=NAME,P003001,P003002,P003003,P003004,P003005,P003006,P003007,P003008&for=block%20group:*&in=state:42&in=county:101&in=tract:*&key={census_key}",
            "fields": [
                "Name",
                "Total",
                "White alone",
                "Black or African American alone",
                "American Indian and Alaska Native alone",
                "Asian alone",
                "Native Hawaiian and Other Pacific Islander alone",
                "Some Other Race alone",
                "Two or More Races",
                "state",
                "county",
                "tract",
                "block_group",
            ],
        },
        {
            "table": "acs_2010",
            "url": "https://api.census.gov/data/2010/acs/acs5?get=NAME,B00002_001E,B19013_001E,B19013_001M&for=tract:*&in=state:42&in=county:101&key={census_key}",
            "fields": [
                "Name",
                "Total Households",
                "Median Household Income",
                "Median Household Income Margin",
                "state",
                "county",
                "tract",
            ],
        },
        {
            "table": "acs_tract_2015",
            "url": "https://api.census.gov/data/2015/acs/acs5?get=NAME,B00002_001E,B19013_001E,B19013_001M&for=tract:*&in=state:42&in=county:101&key={census_key}",
            "fields": [
                "Name",
                "Total Households",
                "Median Household Income",
                "Median Household Income Margin",
                "state",
                "county",
                "tract",
            ],
        },
        {
            "table": "acs_2015",
            "url": "https://api.census.gov/data/2015/acs/acs5?get=NAME,B02001_001E,B02001_002E,B02001_003E,B02001_004E,B02001_005E,B02001_006E,B02001_007E,B02001_008E&for=block%20group:*&in=state:42&in=county:101&in=tract:*&key={census_key}",
            "fields": [
                "Name",
                "Total",
                "White alone",
                "Black or African American alone",
                "American Indian and Alaska Native alone",
                "Asian alone",
                "Native Hawaiian and Other Pacific Islander alone",
                "Some other race alone",
                "Two or more races",
                "state",
                "county",
                "tract",
                "block_group",
            ],
        },
    ]

    for census_config in tqdm(census_configs):
        census(engine, census_schema, **census_config)

    print(f'Process sdp catchment shape files...') 
    catchment_years = [
        '1617',
        '1718',
        '1819',
        '1920',
    ]

    # download the zip file containing all shape files
    sdp_shapes_url = 'https://cdn.philasd.org/offices/performance/Open_Data/School_Information/School_Catchment/SDP_Catchment_All_Years.zip'
    sdp_data_path = os.path.join('..', 'data', 'schools')

    sdp_shape_files_path = download_unzip(sdp_shapes_url, sdp_data_path)

    # process file for each year, if it exists
    for year in tqdm(catchment_years):
        shape_file = unzip_shape_file(year, sdp_shape_files_path)
        local_shapes(file_path=shape_file, schema=sdp_schema, db=db, user=user)

    print("Processing census tract shape files...")
    # TODO consider using the ogr python interface
    # http://pcjericks.github.io/py-gdalogr-cookbook/vector_layers.html#create-a-postgis-table-from-wkt
    print("Processing 2010")
    subprocess.run(
        [
            'ogr2ogr',
            '-f', 'PostgreSQL',
            f'Pg:dbname={db} host=localhost port=5432 user={user}',
            '-lco', f'SCHEMA={census_schema}',
            '-lco', 'OVERWRITE=YES',
            '-nlt', 'PROMOTE_TO_MULTI',
            '-sql', "select 2010 as year, tractce10 as tractce, geoid10 as geoid, name10 as name, aland10 as aland, awater10 as awater from tl_2010_42101_tract10 where countyfp10 = '101'",
            '-t_srs', 'EPSG:2272',
            '-nln', 'tract',
            '/vsizip/vsicurl/https://www2.census.gov/geo/tiger/TIGER2010/TRACT/2010/tl_2010_42101_tract10.zip',
        ],
        check=True,
    )

    print("Processing census block group shape files...")
    print("Processing 2010")
    subprocess.run(
        [
            'ogr2ogr',
            '-f', 'PostgreSQL',
            f'Pg:dbname={db} host=localhost port=5432 user={user}',
            '-lco', f'SCHEMA={census_schema}',
            '-lco', 'OVERWRITE=YES',
            '-nlt', 'PROMOTE_TO_MULTI',
            '-sql', "select 2010 as year, tractce10 as tractce, blkgrpce10 as blkgrpce, geoid10 as geoid, namelsad10 as name, aland10 as aland, awater10 as awater from tl_2010_42101_bg10 where countyfp10 = '101'",
            '-t_srs', 'EPSG:2272',
            '-nln', 'block_group',
            '/vsizip/vsicurl/https://www2.census.gov/geo/tiger/TIGER2010/BG/2010/tl_2010_42101_bg10.zip',
        ],
        check=True,
    )

    engine.dispose()

if __name__ == '__main__':
    main()